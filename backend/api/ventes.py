"""Agrégation des ventes (depuis les commandes payées) et import/template Excel."""
import calendar
import datetime as dt
import io
from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP

from django.db import transaction

from .models import LigneCommande, VenteAgregee, CategoriePlat

COLONNES = ['date', 'categorie', 'montant_ht', 'montant_ttc', 'quantite']


def _periode(annee, mois):
    annee = int(annee)
    if mois:
        mois = int(mois)
        return dt.date(annee, mois, 1), dt.date(annee, mois, calendar.monthrange(annee, mois)[1])
    return dt.date(annee, 1, 1), dt.date(annee, 12, 31)


def _ht(ttc: Decimal, taux: Decimal) -> Decimal:
    return (ttc / (1 + (taux or Decimal('0')) / 100)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


@transaction.atomic
def recalculer_depuis_commandes(annee, mois) -> int:
    """Agrège les commandes PAYÉES de la période. Stocke une ligne par
    (jour, catégorie) + une ligne globale (categorie NULL) par jour.
    Remplace les lignes source='commandes' de la période."""
    debut, fin = _periode(annee, mois)
    lignes = (
        LigneCommande.objects
        .filter(
            commande__created_at__date__gte=debut,
            commande__created_at__date__lte=fin,
            commande__paiement__statut__nom='paye',
        )
        .select_related('commande', 'plat__sous_categorie__categorie')
    )

    par_cat = defaultdict(lambda: {'ht': Decimal('0'), 'ttc': Decimal('0'), 'q': 0})
    par_jour = defaultdict(lambda: {'ht': Decimal('0'), 'ttc': Decimal('0'), 'q': 0})

    for l in lignes:
        jour = l.commande.created_at.date()
        ttc = Decimal(l.quantite) * l.prix_unitaire_snapshot
        ht = _ht(ttc, getattr(l.plat, 'taux_tva', Decimal('0')))
        sc = l.plat.sous_categorie
        cat = sc.categorie if (sc and sc.categorie_id) else None
        if cat:
            k = (jour, cat.id)
            par_cat[k]['ht'] += ht
            par_cat[k]['ttc'] += ttc
            par_cat[k]['q'] += l.quantite
        par_jour[jour]['ht'] += ht
        par_jour[jour]['ttc'] += ttc
        par_jour[jour]['q'] += l.quantite

    VenteAgregee.objects.filter(source='commandes', date__gte=debut, date__lte=fin).delete()

    objs = [
        VenteAgregee(date=j, categorie_id=cid, montant_ht=v['ht'],
                     montant_ttc=v['ttc'], quantite=v['q'], source='commandes')
        for (j, cid), v in par_cat.items()
    ] + [
        VenteAgregee(date=j, categorie=None, montant_ht=v['ht'],
                     montant_ttc=v['ttc'], quantite=v['q'], source='commandes')
        for j, v in par_jour.items()
    ]
    VenteAgregee.objects.bulk_create(objs)
    return len(objs)


def _parse_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    try:
        return dt.datetime.strptime(str(v).strip()[:10], '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def _num(v):
    if v in (None, ''):
        return Decimal('0')
    try:
        return Decimal(str(v).replace(',', '.').strip())
    except (ValueError, TypeError):
        return Decimal('0')


@transaction.atomic
def importer_excel(fichier) -> int:
    """Importe un .xlsx (colonnes date, categorie, montant_ht, montant_ttc,
    quantite). Remplace les lignes source='excel' des dates importées."""
    import openpyxl
    wb = openpyxl.load_workbook(fichier, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError('Fichier vide.')

    header = [str(c).strip().lower() if c is not None else '' for c in rows[0]]
    idx = {name: header.index(name) for name in COLONNES if name in header}
    if 'date' not in idx:
        raise ValueError("Colonne « date » manquante (attendu : %s)." % ', '.join(COLONNES))

    cats = {c.nom.lower(): c for c in CategoriePlat.objects.all()}
    erreurs, objs, dates = [], [], set()

    for n, row in enumerate(rows[1:], start=2):
        def cell(name):
            i = idx.get(name)
            return row[i] if (i is not None and i < len(row)) else None

        dval = cell('date')
        if dval in (None, ''):
            continue
        jour = _parse_date(dval)
        if not jour:
            erreurs.append(f'L{n} : date invalide ({dval})')
            continue

        catname = cell('categorie')
        catname = str(catname).strip() if catname not in (None, '') else ''
        categorie = None
        if catname:
            categorie = cats.get(catname.lower())
            if not categorie:
                erreurs.append(f'L{n} : catégorie inconnue « {catname} »')
                continue

        objs.append(VenteAgregee(
            date=jour, categorie=categorie,
            montant_ht=_num(cell('montant_ht')), montant_ttc=_num(cell('montant_ttc')),
            quantite=int(_num(cell('quantite'))), source='excel',
        ))
        dates.add(jour)

    if erreurs:
        raise ValueError(' ; '.join(erreurs[:15]))
    if not objs:
        raise ValueError('Aucune ligne exploitable.')

    VenteAgregee.objects.filter(source='excel', date__in=dates).delete()
    VenteAgregee.objects.bulk_create(objs)
    return len(objs)


def construire_template() -> bytes:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ventes'
    ws.append(COLONNES)
    ws.append(['2026-06-01', 'Boissons', 1127.27, 1240.50, 310])
    ws.append(['2026-06-01', '', 4090.91, 4500.00, 980])  # ligne globale (catégorie vide)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
